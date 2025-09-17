import os
import requests
import json
import time
import re
from openpyxl import Workbook, load_workbook

# =====================
# CONFIG
# =====================
MODEL_ENDPOINT = "http://localhost:11434/api/generate"
MODEL_NAME = "llama3.1:8b-instruct-q4_K_M"
MAX_RETRIES = 3
TIMEOUT = 60  # seconds


def ask_model(prompt: str) -> str:
    payload = {
        "model": MODEL_NAME,
        "prompt": prompt,
        "max_tokens": 500,
        "temperature": 0.3
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"\n🚀 Sending prompt to model (attempt {attempt})...")
            resp = requests.post(MODEL_ENDPOINT, json=payload, stream=True, timeout=TIMEOUT)
            resp.raise_for_status()

            output_text = ""
            for line in resp.iter_lines():
                if not line:
                    continue
                try:
                    obj = json.loads(line.decode("utf-8"))
                    if "response" in obj:
                        chunk = obj["response"]
                        output_text += chunk
                        print(f"[STREAM] {chunk}", end="", flush=True)
                    if obj.get("done", False):
                        break
                except json.JSONDecodeError:
                    continue

            print("\n\n==== Final Combined Output ====")
            print(output_text.strip())
            print("================================\n")

            return output_text.strip()

        except Exception as e:
            print(f"⚠️ Attempt {attempt} failed: {e}")
            if attempt < MAX_RETRIES:
                wait_time = 2 ** attempt
                print(f"⏳ Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                raise RuntimeError(f"❌ Model request failed after {MAX_RETRIES} attempts: {e}")


def parse_testcase(generated: str):
    """
    Extract fields from model output, handling markdown, bullets, and multiline values.
    """
    fields = {"title": "", "pre": "", "steps": "", "data": "", "expected": ""}

    # Normalize markdown (remove **, *, extra spaces)
    text = re.sub(r"\*\*", "", generated)
    text = re.sub(r"^\s*[-*]\s*", "", text, flags=re.MULTILINE)

    # Regex to capture sections
    pattern = re.compile(
        r"(Title|Pre-Conditions|Test Steps|Test Data|Expected Result)\s*:\s*([\s\S]*?)(?=\n(?:Title|Pre-Conditions|Test Steps|Test Data|Expected Result)\s*:|\Z)",
        re.IGNORECASE
    )

    for match in pattern.finditer(text):
        key = match.group(1).lower()
        value = match.group(2).strip()
        if key.startswith("title"):
            fields["title"] = value
        elif key.startswith("pre"):
            fields["pre"] = value
        elif key.startswith("test steps"):
            fields["steps"] = value
        elif key.startswith("test data"):
            fields["data"] = value
        elif key.startswith("expected"):
            fields["expected"] = value

    return fields


def generate_manual_testcases(req_dir: str, output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    headers = [
        "Test Case ID", "Requirement ID", "Title",
        "Pre-Conditions", "Test Steps", "Test Data",
        "Expected Result", "Actual Result", "Status", "Remarks"
    ]
    ws.append(headers)

    if not os.path.exists(req_dir):
        raise FileNotFoundError(f"❌ Requirements directory not found: {req_dir}")

    req_files = sorted([f for f in os.listdir(req_dir) if f.endswith(".txt")])
    if not req_files:
        print(f"⚠️ No .txt files found in {req_dir}")
        return

    tc_count = 1

    for req_file in req_files:
        with open(os.path.join(req_dir, req_file), "r", encoding="utf-8") as f:
            req_text = f.read()

        prompt = f"""
You are a QA engineer. Convert this requirement into a manual test case.
Requirement:
{req_text}

Output format (strict, no extra explanation):
Title:
Pre-Conditions:
Test Steps:
Test Data:
Expected Result:
"""
        generated = ask_model(prompt)
        fields = parse_testcase(generated)

        row_data = [
            f"TC-{tc_count:03}",
            req_file.replace(".txt", ""),
            fields["title"],
            fields["pre"],
            fields["steps"],
            fields["data"],
            fields["expected"],
            "", "", ""
        ]

        ws.append(row_data)
        print(f"✅ Added Test Case {tc_count}: {row_data[:3]}")
        tc_count += 1

    wb.save(output_file)
    print(f"\n📊 Total {tc_count-1} test cases written to {output_file}")

    # Verify last few rows
    wb_check = load_workbook(output_file)
    ws_check = wb_check.active
    print("\n🔎 Preview of last few rows in Excel:")
    for row in ws_check.iter_rows(min_row=max(2, ws_check.max_row-4), max_row=ws_check.max_row, values_only=True):
        print(row)


if __name__ == "__main__":
    req_dir = "clientA-data/text/normalized" 
    output_file = "clientA-data/train/Manual_TestCases.xlsx"
    generate_manual_testcases(req_dir, output_file)
